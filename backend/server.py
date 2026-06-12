"""Passover Seating Manager — FastAPI backend.

Phase 1 (intake form) + Phase 2 (staff admin dashboard).
Backed by Supabase Postgres via SQLAlchemy + asyncpg.
"""
from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path(__file__).parent / ".env")

import os
import bcrypt
import jwt
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from rapidfuzz import fuzz, process as fuzz_process

from db import engine, AsyncSessionLocal, init_db, get_db

JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALG = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 12

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("passover")

app = FastAPI(title="Passover Seating API")
api = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)


# ---------- helpers ----------
def hash_pw(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_pw(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def make_token(user_id: int, username: str, is_admin: bool) -> str:
    payload = {
        "sub": str(user_id), "username": username, "is_admin": is_admin,
        "exp": datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(
    creds: Optional[HTTPAuthorizationCredentials] = Depends(security),
    db: AsyncSession = Depends(get_db),
):
    if not creds:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    row = (await db.execute(
        text("SELECT id, username, display_name, is_admin, is_active FROM staff_users WHERE id=:id"),
        {"id": int(payload["sub"])},
    )).mappings().first()
    if not row or not row["is_active"]:
        raise HTTPException(401, "User inactive or missing")
    return dict(row)


async def require_admin(user: dict = Depends(get_current_user)):
    if not user["is_admin"]:
        raise HTTPException(403, "Admin required")
    return user


async def log_activity(db: AsyncSession, user: dict, action_type: str, guest_id=None, table_id=None, ballroom_id=None, details=None):
    import json as _json
    await db.execute(text("""
        INSERT INTO activity_log (action_type, staff_member_name, staff_user_id, guest_id, table_id, ballroom_id, details)
        VALUES (:a, :n, :u, :g, :t, :b, CAST(:d AS jsonb))
    """), {"a": action_type, "n": user["display_name"], "u": user["id"],
           "g": guest_id, "t": table_id, "b": ballroom_id, "d": _json.dumps(details or {})})


# ---------- pydantic models ----------
class GuestInput(BaseModel):
    fullName: str = Field(min_length=1)
    invoiceNumber: str = Field(min_length=1)
    partySize: int = Field(ge=1)
    seatingPreferences: List[str] = Field(default_factory=list, max_length=5)
    # Optional parallel list of linked roster invoice numbers (one per preference, "" if not picked from autocomplete)
    linkedInvoiceNumbers: List[str] = Field(default_factory=list, max_length=5)
    highChairNeeded: bool = False
    highChairCount: int = 0
    specialNotes: Optional[str] = None

class LoginInput(BaseModel):
    username: str
    password: str

class StaffCreateInput(BaseModel):
    username: str
    password: str
    displayName: str
    isAdmin: bool = False

class StaffUpdateInput(BaseModel):
    displayName: Optional[str] = None
    isAdmin: Optional[bool] = None
    isActive: Optional[bool] = None
    password: Optional[str] = None

class StaffNoteInput(BaseModel):
    note: str = Field(min_length=1)

class PreferenceResolveInput(BaseModel):
    resolutionStatus: str  # 'confirmed' | 'no_match'
    resolvedGuestId: Optional[int] = None

class GuestUpdateInput(BaseModel):
    fullName: Optional[str] = None
    partySize: Optional[int] = None
    highChairNeeded: Optional[bool] = None
    highChairCount: Optional[int] = None
    specialNotes: Optional[str] = None
    isDuplicate: Optional[bool] = None
    familyId: Optional[str] = None
    nearFamilyId: Optional[str] = None
    tableId: Optional[int] = None
    invoiceNumber: Optional[str] = None
    seatingPreferences: Optional[List[str]] = None


# ---------- serializers ----------
def guest_to_api(g) -> dict:
    return {
        "id": g["id"], "fullName": g["full_name"], "invoiceNumber": g["invoice_number"],
        "partySize": g["party_size"], "seatingPreferences": list(g["seating_preferences"] or []),
        "highChairNeeded": g["high_chair_needed"], "highChairCount": g["high_chair_count"],
        "status": g["status"], "ballroomId": g["ballroom_id"], "tableId": g["table_id"],
        "specialNotes": g["special_notes"], "isDuplicate": g["is_duplicate"],
        "familyId": g["family_id"] if "family_id" in g.keys() else None,
        "nearFamilyId": g["near_family_id"] if "near_family_id" in g.keys() else None,
        "submissionTimestamp": g["submission_timestamp"].isoformat(),
        "lastUpdatedTimestamp": g["last_updated_timestamp"].isoformat(),
    }

def note_to_api(n) -> dict:
    return {"id": n["id"], "guestId": n["guest_id"], "note": n["note"],
            "staffName": n["staff_name"], "createdAt": n["created_at"].isoformat()}

def pref_to_api(p) -> dict:
    return {"id": p["id"], "guestId": p["guest_id"], "preferenceIndex": p["preference_index"],
            "preferenceName": p["preference_name"], "resolutionStatus": p["resolution_status"],
            "resolvedGuestId": p["resolved_guest_id"],
            "fuzzyScore": float(p["fuzzy_score"]) if p["fuzzy_score"] is not None else None,
            "resolvedAt": p["resolved_at"].isoformat() if p["resolved_at"] else None,
            "createdAt": p["created_at"].isoformat()}


# ---------- PUBLIC ENDPOINTS ----------
@api.get("/health")
async def health():
    return {"status": "ok"}


@api.post("/guests", status_code=201)
async def submit_guest(body: GuestInput, db: AsyncSession = Depends(get_db)):
    """Phase 1 intake form submission. Public, no auth."""
    existing = (await db.execute(
        text("SELECT id FROM guests WHERE invoice_number=:inv"),
        {"inv": body.invoiceNumber},
    )).fetchall()
    is_dup = len(existing) > 0

    prefs = [p for p in (body.seatingPreferences or []) if p.strip()]
    linked = list(body.linkedInvoiceNumbers or [])
    while len(linked) < len(prefs): linked.append("")

    res = (await db.execute(text("""
        INSERT INTO guests (full_name, invoice_number, party_size, seating_preferences,
                            high_chair_needed, high_chair_count, special_notes, is_duplicate)
        VALUES (:fn, :inv, :ps, :prefs, :hcn, :hcc, :sn, :dup)
        RETURNING *
    """), {"fn": body.fullName, "inv": body.invoiceNumber, "ps": body.partySize,
            "prefs": prefs, "hcn": body.highChairNeeded,
            "hcc": body.highChairCount if body.highChairNeeded else 0,
            "sn": body.specialNotes, "dup": is_dup})).mappings().first()

    for idx, name in enumerate(prefs):
        link_inv = (linked[idx] or "").strip() or None
        # if linked to a roster invoice, see if that guest has already submitted — auto-confirm
        resolved_gid = None; status = "pending"; score = None
        if link_inv:
            other = (await db.execute(text(
                "SELECT id FROM guests WHERE invoice_number=:i AND id<>:self ORDER BY submission_timestamp ASC LIMIT 1"
            ), {"i": link_inv, "self": res["id"]})).mappings().first()
            if other:
                resolved_gid = other["id"]; status = "confirmed"; score = 1.0
        await db.execute(text("""
            INSERT INTO preference_resolutions (guest_id, preference_index, preference_name,
                linked_invoice_number, resolved_guest_id, resolution_status, fuzzy_score, resolved_at)
            VALUES (:g, :i, :n, :li, :rg, CAST(:s AS preference_resolution_status), :sc,
                    CASE WHEN :s='confirmed' THEN NOW() ELSE NULL END)
        """), {"g": res["id"], "i": idx, "n": name, "li": link_inv,
               "rg": resolved_gid, "s": status, "sc": score})

    # Auto-link: any pending pref_resolutions whose linked_invoice_number = this new guest's invoice
    # should now be confirmed pointing at this new guest
    await db.execute(text("""
        UPDATE preference_resolutions
        SET resolved_guest_id=:gid,
            resolution_status='confirmed',
            fuzzy_score=1.0,
            resolved_at=NOW()
        WHERE linked_invoice_number=:inv
          AND resolution_status='pending'
          AND guest_id<>:gid
    """), {"gid": res["id"], "inv": body.invoiceNumber})

    await db.commit()
    return {"guest": guest_to_api(res), "isDuplicate": is_dup, "priorSubmissionCount": len(existing)}


@api.get("/guests/check-invoice/{invoice_number}")
async def check_invoice(invoice_number: str, db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(
        text("SELECT id FROM guests WHERE invoice_number=:i"), {"i": invoice_number},
    )).fetchall()
    return {"invoiceNumber": invoice_number, "hasSubmissions": len(rows) > 0, "submissionCount": len(rows)}


# ---------- REGISTERED GUESTS (master roster) ----------
class RegisteredGuestInput(BaseModel):
    invoiceNumber: str = Field(min_length=1)
    fullName: str = Field(min_length=1)
    email: Optional[str] = None
    phone: Optional[str] = None
    notes: Optional[str] = None


def rg_to_api(r) -> dict:
    return {"id": r["id"], "invoiceNumber": r["invoice_number"], "fullName": r["full_name"],
            "email": r["email"], "phone": r["phone"], "notes": r["notes"],
            "createdAt": r["created_at"].isoformat()}


@api.get("/roster/lookup/{invoice_number}")
async def roster_lookup(invoice_number: str, db: AsyncSession = Depends(get_db)):
    """Public — used by intake form to auto-fill name when an invoice number is entered."""
    r = (await db.execute(text(
        "SELECT * FROM registered_guests WHERE invoice_number=:i"
    ), {"i": invoice_number.strip()})).mappings().first()
    if not r:
        return {"found": False}
    return {"found": True, "fullName": r["full_name"], "invoiceNumber": r["invoice_number"]}


@api.get("/roster/search")
async def roster_search(q: str = "", excludeInvoice: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Public — used by intake form's seating-preference autocomplete."""
    q = (q or "").strip()
    if len(q) < 1:
        return []
    params = {"q": f"%{q}%", "limit": 12}
    extra = ""
    if excludeInvoice:
        extra = "AND invoice_number<>:exc"
        params["exc"] = excludeInvoice
    rows = (await db.execute(text(f"""
        SELECT id, invoice_number, full_name FROM registered_guests
        WHERE (full_name ILIKE :q OR invoice_number ILIKE :q) {extra}
        ORDER BY full_name ASC LIMIT :limit
    """), params)).mappings().all()
    return [{"id": r["id"], "invoiceNumber": r["invoice_number"], "fullName": r["full_name"]} for r in rows]


@api.get("/roster")
async def roster_list(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
                     search: Optional[str] = None):
    where, params = "", {}
    if search:
        where = "WHERE full_name ILIKE :s OR invoice_number ILIKE :s"
        params["s"] = f"%{search}%"
    rows = (await db.execute(text(
        f"SELECT * FROM registered_guests {where} ORDER BY full_name ASC LIMIT 2000"
    ), params)).mappings().all()
    return [rg_to_api(r) for r in rows]


@api.post("/roster", status_code=201)
async def roster_create(body: RegisteredGuestInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    try:
        row = (await db.execute(text("""
            INSERT INTO registered_guests (invoice_number, full_name, email, phone, notes)
            VALUES (:i, :n, :e, :p, :nt) RETURNING *
        """), {"i": body.invoiceNumber.strip(), "n": body.fullName.strip(),
               "e": body.email, "p": body.phone, "nt": body.notes})).mappings().first()
        await log_activity(db, user, "roster_create", details={"invoice": body.invoiceNumber})
        await db.commit()
        return rg_to_api(row)
    except Exception as e:
        await db.rollback()
        if "duplicate" in str(e).lower():
            raise HTTPException(400, "Invoice number already in roster")
        raise HTTPException(400, str(e)[:200])


@api.delete("/roster/{rg_id}", status_code=204)
async def roster_delete(rg_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    await db.execute(text("DELETE FROM registered_guests WHERE id=:i"), {"i": rg_id})
    await log_activity(db, user, "roster_delete", details={"id": rg_id})
    await db.commit()


@app.post("/api/roster/import-csv")
async def roster_import_csv(
    request: Request,
    user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db),
):
    """Accepts multipart upload with a 'file' field, OR raw CSV body with text/csv content-type.
    CSV columns (header row required, case-insensitive): invoice_number (or 'invoice'), full_name (or 'name').
    Optional: email, phone, notes. Updates on conflict by invoice_number."""
    import csv as _csv, io
    from fastapi import UploadFile
    ctype = (request.headers.get("content-type") or "").lower()
    content = ""
    if "multipart" in ctype:
        form = await request.form()
        f = form.get("file")
        if f is None:
            raise HTTPException(400, "Missing 'file' field")
        content = (await f.read()).decode("utf-8-sig", errors="replace")
    else:
        content = (await request.body()).decode("utf-8-sig", errors="replace")

    if not content.strip():
        raise HTTPException(400, "Empty CSV body")

    reader = _csv.DictReader(io.StringIO(content))
    norm = lambda s: (s or "").strip().lower().replace(" ", "_")
    if not reader.fieldnames:
        raise HTTPException(400, "CSV missing header row")
    field_map = {norm(f): f for f in reader.fieldnames}
    inv_col = field_map.get("invoice_number") or field_map.get("invoice") or field_map.get("invoiceno") or field_map.get("invoice#")
    name_col = field_map.get("full_name") or field_map.get("name") or field_map.get("guest_name") or field_map.get("customer")
    if not inv_col or not name_col:
        raise HTTPException(400, f"CSV must include an invoice and a name column. Detected headers: {list(reader.fieldnames)}")
    email_col = field_map.get("email")
    phone_col = field_map.get("phone") or field_map.get("phone_number")
    notes_col = field_map.get("notes") or field_map.get("note")

    inserted = updated = skipped = 0
    errors = []
    for i, row in enumerate(reader, start=2):
        inv = (row.get(inv_col) or "").strip()
        nm = (row.get(name_col) or "").strip()
        if not inv or not nm:
            skipped += 1; errors.append(f"Row {i}: missing invoice or name"); continue
        try:
            existing = (await db.execute(text(
                "SELECT id FROM registered_guests WHERE invoice_number=:i"
            ), {"i": inv})).fetchone()
            if existing:
                await db.execute(text("""
                    UPDATE registered_guests SET full_name=:n,
                        email=COALESCE(:e, email), phone=COALESCE(:p, phone), notes=COALESCE(:nt, notes)
                    WHERE invoice_number=:i
                """), {"i": inv, "n": nm,
                       "e": row.get(email_col) if email_col else None,
                       "p": row.get(phone_col) if phone_col else None,
                       "nt": row.get(notes_col) if notes_col else None})
                updated += 1
            else:
                await db.execute(text("""
                    INSERT INTO registered_guests (invoice_number, full_name, email, phone, notes)
                    VALUES (:i, :n, :e, :p, :nt)
                """), {"i": inv, "n": nm,
                       "e": row.get(email_col) if email_col else None,
                       "p": row.get(phone_col) if phone_col else None,
                       "nt": row.get(notes_col) if notes_col else None})
                inserted += 1
        except Exception as ex:
            skipped += 1; errors.append(f"Row {i}: {str(ex)[:120]}")

    await log_activity(db, user, "roster_import", details={"inserted": inserted, "updated": updated, "skipped": skipped})
    await db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors[:20]}


# ---------- AUTH ----------
@api.post("/auth/login")
async def login(body: LoginInput, db: AsyncSession = Depends(get_db)):
    row = (await db.execute(
        text("SELECT * FROM staff_users WHERE username=:u"), {"u": body.username},
    )).mappings().first()
    if not row or not row["is_active"] or not verify_pw(body.password, row["password_hash"]):
        raise HTTPException(401, "Invalid username or password")
    await db.execute(text("UPDATE staff_users SET last_login=NOW() WHERE id=:i"), {"i": row["id"]})
    await log_activity(db, dict(row), "login")
    await db.commit()
    token = make_token(row["id"], row["username"], row["is_admin"])
    return {"token": token, "user": {"id": row["id"], "username": row["username"],
            "displayName": row["display_name"], "isAdmin": row["is_admin"]}}


@api.post("/auth/logout")
async def logout(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    await log_activity(db, user, "logout")
    await db.commit()
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return {"id": user["id"], "username": user["username"],
            "displayName": user["display_name"], "isAdmin": user["is_admin"]}


# ---------- STAFF MANAGEMENT (admin) ----------
@api.get("/staff")
async def list_staff(user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text(
        "SELECT id, username, display_name, is_admin, is_active, created_at, last_login FROM staff_users ORDER BY id"
    ))).mappings().all()
    return [{"id": r["id"], "username": r["username"], "displayName": r["display_name"],
             "isAdmin": r["is_admin"], "isActive": r["is_active"],
             "createdAt": r["created_at"].isoformat(),
             "lastLogin": r["last_login"].isoformat() if r["last_login"] else None} for r in rows]


@api.post("/staff", status_code=201)
async def create_staff(body: StaffCreateInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    try:
        row = (await db.execute(text("""
            INSERT INTO staff_users (username, password_hash, display_name, is_admin)
            VALUES (:u, :p, :d, :a) RETURNING id, username, display_name, is_admin, is_active
        """), {"u": body.username, "p": hash_pw(body.password), "d": body.displayName, "a": body.isAdmin})).mappings().first()
        await log_activity(db, user, "staff_create", details={"newUserId": row["id"], "username": body.username})
        await db.commit()
        return {"id": row["id"], "username": row["username"], "displayName": row["display_name"],
                "isAdmin": row["is_admin"], "isActive": row["is_active"]}
    except Exception as e:
        await db.rollback()
        raise HTTPException(400, f"Could not create staff: {str(e)[:200]}")


@api.patch("/staff/{staff_id}")
async def update_staff(staff_id: int, body: StaffUpdateInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    sets = []
    params: dict = {"id": staff_id}
    if body.displayName is not None: sets.append("display_name=:dn"); params["dn"] = body.displayName
    if body.isAdmin is not None: sets.append("is_admin=:ia"); params["ia"] = body.isAdmin
    if body.isActive is not None: sets.append("is_active=:ic"); params["ic"] = body.isActive
    if body.password is not None: sets.append("password_hash=:ph"); params["ph"] = hash_pw(body.password)
    if not sets:
        raise HTTPException(400, "No fields to update")
    await db.execute(text(f"UPDATE staff_users SET {', '.join(sets)} WHERE id=:id"), params)
    await log_activity(db, user, "staff_update", details={"targetId": staff_id, "fields": list(params.keys())})
    await db.commit()
    return {"ok": True}


# ---------- GUESTS (staff) ----------
@api.get("/guests")
async def list_guests(
    user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
    search: Optional[str] = None, status: Optional[str] = None,
    ballroomId: Optional[int] = None, isDuplicate: Optional[bool] = None,
    highChair: Optional[bool] = None, hasNotes: Optional[bool] = None,
    sort: Optional[str] = "submissionTimestamp", order: Optional[str] = "desc",
):
    conds = []; params: dict = {}
    if search:
        conds.append("(full_name ILIKE :s OR invoice_number ILIKE :s)"); params["s"] = f"%{search}%"
    if status:
        conds.append("status = :st"); params["st"] = status
    if ballroomId is not None:
        conds.append("ballroom_id = :br"); params["br"] = ballroomId
    if isDuplicate is not None:
        conds.append("is_duplicate = :dup"); params["dup"] = isDuplicate
    if highChair is not None:
        conds.append("high_chair_needed = :hc"); params["hc"] = highChair
    sort_col = {"submissionTimestamp": "submission_timestamp", "fullName": "full_name",
                "partySize": "party_size", "invoiceNumber": "invoice_number"}.get(sort, "submission_timestamp")
    order_sql = "DESC" if (order or "desc").lower() == "desc" else "ASC"
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows = (await db.execute(text(f"SELECT * FROM guests {where} ORDER BY {sort_col} {order_sql} LIMIT 5000"), params)).mappings().all()
    result = [guest_to_api(r) for r in rows]
    if hasNotes is not None:
        ids = [g["id"] for g in result]
        if ids:
            note_rows = (await db.execute(text("SELECT DISTINCT guest_id FROM staff_notes WHERE guest_id = ANY(:ids)"), {"ids": ids})).fetchall()
            with_notes = {r[0] for r in note_rows}
            result = [g for g in result if (g["id"] in with_notes) == hasNotes]
    return result


@api.get("/guests/unassigned")
async def unassigned_queue(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text(
        "SELECT * FROM guests WHERE status='unassigned' ORDER BY submission_timestamp ASC"
    ))).mappings().all()
    return [guest_to_api(r) for r in rows]


@api.get("/guests/stats")
async def stats(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    g = (await db.execute(text("""
      SELECT COUNT(*) AS total,
             COALESCE(SUM(party_size),0) AS total_people,
             COUNT(*) FILTER (WHERE is_duplicate) AS dups,
             COUNT(*) FILTER (WHERE status='unassigned') AS unassigned,
             COUNT(*) FILTER (WHERE status='partially_assigned') AS partial,
             COUNT(*) FILTER (WHERE status='fully_assigned') AS full,
             COALESCE(SUM(high_chair_count),0) AS high_chairs
      FROM guests
    """))).mappings().first()
    unresolved = (await db.execute(text(
        "SELECT COUNT(*) AS c FROM preference_resolutions WHERE resolution_status IN ('pending','auto_suggested')"
    ))).scalar()
    total = g["total"] or 0
    seated_pct = (g["full"] / total * 100) if total else 0
    unassigned_pct = (g["unassigned"] / total * 100) if total else 0
    return {
        "totalSubmissions": total, "totalPeople": int(g["total_people"]),
        "totalDuplicates": g["dups"], "totalHighChairs": int(g["high_chairs"]),
        "unresolvedPreferences": unresolved or 0,
        "percentSeated": round(seated_pct, 1), "percentUnassigned": round(unassigned_pct, 1),
        "statusBreakdown": {"unassigned": g["unassigned"], "partially_assigned": g["partial"], "fully_assigned": g["full"]},
    }


@api.get("/guests/{guest_id}")
async def get_guest(guest_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("SELECT * FROM guests WHERE id=:i"), {"i": guest_id})).mappings().first()
    if not row: raise HTTPException(404, "Guest not found")
    return guest_to_api(row)


@api.patch("/guests/{guest_id}")
async def update_guest(guest_id: int, body: GuestUpdateInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    before = (await db.execute(text("SELECT * FROM guests WHERE id=:i"), {"i": guest_id})).mappings().first()
    if not before: raise HTTPException(404, "Guest not found")
    sets, params = [], {"id": guest_id}
    mapping = {"fullName": "full_name", "partySize": "party_size", "highChairNeeded": "high_chair_needed",
               "highChairCount": "high_chair_count", "specialNotes": "special_notes", "isDuplicate": "is_duplicate",
               "familyId": "family_id", "nearFamilyId": "near_family_id",
               "invoiceNumber": "invoice_number"}
    for k, col in mapping.items():
        v = getattr(body, k)
        if v is not None: sets.append(f"{col}=:{col}"); params[col] = v
    if body.seatingPreferences is not None:
        sets.append("seating_preferences=:sprefs"); params["sprefs"] = list(body.seatingPreferences)[:5]
    if body.tableId is not None:
        if body.tableId == 0:
            sets.append("table_id=NULL"); sets.append("ballroom_id=NULL"); sets.append("status='unassigned'")
        else:
            t = (await db.execute(text("SELECT ballroom_id FROM tables WHERE id=:i"), {"i": body.tableId})).mappings().first()
            if not t: raise HTTPException(404, "Target table not found")
            sets.append("table_id=:tid"); sets.append("ballroom_id=:brid"); sets.append("status='fully_assigned'")
            params["tid"] = body.tableId; params["brid"] = t["ballroom_id"]
    if not sets: raise HTTPException(400, "No fields to update")
    sets.append("last_updated_timestamp=NOW()")
    row = (await db.execute(text(f"UPDATE guests SET {', '.join(sets)} WHERE id=:id RETURNING *"), params)).mappings().first()
    await log_activity(db, user, "guest_update", guest_id=guest_id, details=body.model_dump(exclude_none=True))
    await record_history(db, user, "guests", "guest_update", "guest", str(guest_id),
                         _guest_snap(before), _guest_snap(row))
    await db.commit()
    return guest_to_api(row)


# ---------- STAFF NOTES ----------
@api.get("/guests/{guest_id}/notes")
async def list_notes(guest_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text(
        "SELECT * FROM staff_notes WHERE guest_id=:g ORDER BY created_at DESC"
    ), {"g": guest_id})).mappings().all()
    return [note_to_api(r) for r in rows]


@api.post("/guests/{guest_id}/notes", status_code=201)
async def add_note(guest_id: int, body: StaffNoteInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("""
        INSERT INTO staff_notes (guest_id, note, staff_name, staff_user_id)
        VALUES (:g, :n, :sn, :su) RETURNING *
    """), {"g": guest_id, "n": body.note, "sn": user["display_name"], "su": user["id"]})).mappings().first()
    await log_activity(db, user, "note_add", guest_id=guest_id, details={"noteId": row["id"]})
    await db.commit()
    return note_to_api(row)


@api.delete("/guests/{guest_id}/notes/{note_id}", status_code=204)
async def delete_note(guest_id: int, note_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    await db.execute(text("DELETE FROM staff_notes WHERE id=:i AND guest_id=:g"), {"i": note_id, "g": guest_id})
    await log_activity(db, user, "note_delete", guest_id=guest_id, details={"noteId": note_id})
    await db.commit()


# ---------- PREFERENCES ----------
@api.get("/preferences/mutual")
async def mutual_prefs(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Pairs where guest A's confirmed preference resolves to guest B AND B's confirmed preference resolves to A."""
    rows = (await db.execute(text("""
        SELECT a.guest_id AS a_id, a.resolved_guest_id AS b_id,
               ga.full_name AS a_name, gb.full_name AS b_name,
               ga.party_size AS a_size, gb.party_size AS b_size
        FROM preference_resolutions a
        JOIN preference_resolutions b ON b.guest_id = a.resolved_guest_id
                                       AND b.resolved_guest_id = a.guest_id
                                       AND b.resolution_status='confirmed'
        JOIN guests ga ON ga.id = a.guest_id
        JOIN guests gb ON gb.id = a.resolved_guest_id
        WHERE a.resolution_status='confirmed' AND a.guest_id < a.resolved_guest_id
    """))).mappings().all()
    return [dict(r) for r in rows]


@api.get("/preferences/one-way")
async def one_way_prefs(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Confirmed A→B where B doesn't have a confirmed reverse link to A."""
    rows = (await db.execute(text("""
        SELECT a.id, a.guest_id, a.resolved_guest_id,
               ga.full_name AS a_name, gb.full_name AS b_name
        FROM preference_resolutions a
        JOIN guests ga ON ga.id = a.guest_id
        JOIN guests gb ON gb.id = a.resolved_guest_id
        WHERE a.resolution_status='confirmed'
          AND NOT EXISTS (
            SELECT 1 FROM preference_resolutions b
            WHERE b.guest_id = a.resolved_guest_id AND b.resolved_guest_id = a.guest_id
              AND b.resolution_status='confirmed'
          )
    """))).mappings().all()
    return [dict(r) for r in rows]


@api.get("/preferences/unresolved")
async def unresolved_prefs(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Pending preferences with fuzzy-match suggestions against all guest names."""
    rows = (await db.execute(text("""
        SELECT p.*, g.full_name AS requester_name
        FROM preference_resolutions p
        JOIN guests g ON g.id = p.guest_id
        WHERE p.resolution_status IN ('pending','auto_suggested')
        ORDER BY p.created_at ASC
        LIMIT 500
    """))).mappings().all()
    guests = (await db.execute(text("SELECT id, full_name FROM guests"))).mappings().all()
    names = [(g["id"], g["full_name"]) for g in guests]
    result = []
    for r in rows:
        choices = {fid: name for fid, name in names if fid != r["guest_id"]}
        matches = []
        if choices:
            top = fuzz_process.extract(r["preference_name"], choices, scorer=fuzz.WRatio, limit=3)
            for matched_name, score, gid in top:
                if score >= 60:
                    matches.append({"guestId": gid, "name": matched_name, "score": round(score / 100, 4)})
        d = pref_to_api(r)
        d["requesterName"] = r["requester_name"]
        d["suggestions"] = matches
        result.append(d)
    return result


@api.patch("/preferences/{pref_id}/resolve")
async def resolve_pref(pref_id: int, body: PreferenceResolveInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if body.resolutionStatus not in ("confirmed", "no_match"):
        raise HTTPException(400, "Invalid status")
    score = None
    if body.resolutionStatus == "confirmed" and body.resolvedGuestId:
        pref = (await db.execute(text("SELECT preference_name FROM preference_resolutions WHERE id=:i"), {"i": pref_id})).mappings().first()
        guest = (await db.execute(text("SELECT full_name FROM guests WHERE id=:i"), {"i": body.resolvedGuestId})).mappings().first()
        if pref and guest:
            score = fuzz.WRatio(pref["preference_name"], guest["full_name"]) / 100
    row = (await db.execute(text("""
        UPDATE preference_resolutions
        SET resolution_status = CAST(:s AS preference_resolution_status),
            resolved_guest_id = :rg, resolved_at = NOW(),
            fuzzy_score = :fs
        WHERE id = :i RETURNING *
    """), {"s": body.resolutionStatus, "rg": body.resolvedGuestId, "fs": score, "i": pref_id})).mappings().first()
    if not row: raise HTTPException(404, "Resolution not found")
    await log_activity(db, user, "preference_resolve", guest_id=row["guest_id"],
                        details={"resolutionId": pref_id, "status": body.resolutionStatus, "resolvedGuestId": body.resolvedGuestId})
    await db.commit()
    return pref_to_api(row)


@api.get("/guests/{guest_id}/preference-resolutions")
async def guest_prefs(guest_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text(
        "SELECT * FROM preference_resolutions WHERE guest_id=:g ORDER BY preference_index"
    ), {"g": guest_id})).mappings().all()
    return [pref_to_api(r) for r in rows]


# ---------- BALLROOMS / TABLES / SEATING ----------
class BallroomInput(BaseModel):
    name: str
    widthFt: Optional[float] = None
    heightFt: Optional[float] = None


class TableInput(BaseModel):
    tableNumber: int
    label: Optional[str] = None
    ballroomId: int
    shape: str = "round"  # round | rectangular | square
    maxCapacity: int = 10
    canvasX: Optional[float] = 0
    canvasY: Optional[float] = 0
    rotation: Optional[float] = 0
    widthIn: Optional[float] = None
    lengthIn: Optional[float] = None
    notes: Optional[str] = None
    typeId: Optional[int] = None
    groupId: Optional[str] = None


class TableUpdateInput(BaseModel):
    tableNumber: Optional[int] = None
    label: Optional[str] = None
    ballroomId: Optional[int] = None
    shape: Optional[str] = None
    maxCapacity: Optional[int] = None
    canvasX: Optional[float] = None
    canvasY: Optional[float] = None
    rotation: Optional[float] = None
    widthIn: Optional[float] = None
    lengthIn: Optional[float] = None
    notes: Optional[str] = None
    typeId: Optional[int] = None
    groupId: Optional[str] = None


class AssignInput(BaseModel):
    guestId: int
    allowOverflow: bool = False  # override capacity check


class SeatedInput(BaseModel):
    seated: bool


def table_color(seats_taken: int, capacity: int) -> str:
    if seats_taken == 0: return "gray"
    if seats_taken >= capacity: return "green"
    if capacity - seats_taken <= 2: return "yellow"
    return "blue"


def ballroom_to_api(b) -> dict:
    return {"id": b["id"], "name": b["name"],
            "widthFt": float(b["width_ft"]) if b["width_ft"] is not None else None,
            "heightFt": float(b["height_ft"]) if b["height_ft"] is not None else None,
            "backgroundImageUrl": b["background_image_url"],
            "scaleFactor": float(b["scale_factor"]) if b["scale_factor"] is not None else 1.0,
            "snapEnabled": bool(b["snap_enabled"]) if "snap_enabled" in b.keys() and b["snap_enabled"] is not None else True,
            "gridSizeIn": float(b["grid_size_in"]) if "grid_size_in" in b.keys() and b["grid_size_in"] is not None else 6.0,
            "bgOpacity": float(b["bg_opacity"]) if "bg_opacity" in b.keys() and b["bg_opacity"] is not None else 0.55,
            "bgVisible": bool(b["bg_visible"]) if "bg_visible" in b.keys() and b["bg_visible"] is not None else True,
            "bgCalibration": (b["bg_calibration"] if "bg_calibration" in b.keys() and b["bg_calibration"] is not None else {}),
            "pxPerFt": float(b["px_per_ft"]) if "px_per_ft" in b.keys() and b["px_per_ft"] is not None else 12.0,
            "createdAt": b["created_at"].isoformat()}


def table_to_api(t, seated=0) -> dict:
    cap = t["max_capacity"]
    return {"id": t["id"], "tableNumber": t["table_number"], "label": t["label"],
            "ballroomId": t["ballroom_id"], "shape": t["shape"],
            "maxCapacity": cap, "seatsTaken": seated, "seatsRemaining": cap - seated,
            "color": table_color(seated, cap),
            "canvasX": float(t["canvas_x"]) if t["canvas_x"] is not None else 0,
            "canvasY": float(t["canvas_y"]) if t["canvas_y"] is not None else 0,
            "rotation": float(t["rotation"]) if t["rotation"] is not None else 0,
            "widthIn": float(t["width_in"]) if "width_in" in t.keys() and t["width_in"] is not None else 60.0,
            "lengthIn": float(t["length_in"]) if "length_in" in t.keys() and t["length_in"] is not None else 60.0,
            "typeId": t["type_id"] if "type_id" in t.keys() else None,
            "groupId": t["group_id"] if "group_id" in t.keys() else None,
            "notes": t["notes"]}


@api.get("/ballrooms")
async def list_ballrooms(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text("SELECT * FROM ballrooms ORDER BY name"))).mappings().all()
    return [ballroom_to_api(r) for r in rows]


@api.post("/ballrooms", status_code=201)
async def create_ballroom(body: BallroomInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("""
        INSERT INTO ballrooms (name, width_ft, height_ft) VALUES (:n, :w, :h) RETURNING *
    """), {"n": body.name, "w": body.widthFt, "h": body.heightFt})).mappings().first()
    await log_activity(db, user, "ballroom_create", ballroom_id=row["id"], details={"name": body.name})
    await db.commit()
    return ballroom_to_api(row)


@api.patch("/ballrooms/{ballroom_id}")
async def update_ballroom(ballroom_id: int, body: BallroomInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("""
        UPDATE ballrooms SET name=:n, width_ft=:w, height_ft=:h WHERE id=:i RETURNING *
    """), {"i": ballroom_id, "n": body.name, "w": body.widthFt, "h": body.heightFt})).mappings().first()
    if not row: raise HTTPException(404, "Ballroom not found")
    await log_activity(db, user, "ballroom_update", ballroom_id=ballroom_id)
    await db.commit()
    return ballroom_to_api(row)


@api.delete("/ballrooms/{ballroom_id}", status_code=204)
async def delete_ballroom(ballroom_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    n = (await db.execute(text("SELECT COUNT(*) FROM tables WHERE ballroom_id=:i"), {"i": ballroom_id})).scalar()
    if n: raise HTTPException(400, f"Ballroom has {n} tables — move or delete them first")
    await db.execute(text("DELETE FROM ballrooms WHERE id=:i"), {"i": ballroom_id})
    await log_activity(db, user, "ballroom_delete", ballroom_id=ballroom_id)
    await db.commit()


class FloorPlanInput(BaseModel):
    backgroundImageUrl: Optional[str] = None  # data URL or http URL
    scaleFactor: Optional[float] = None


class CanvasSettingsInput(BaseModel):
    snapEnabled: Optional[bool] = None
    gridSizeIn: Optional[float] = None
    bgOpacity: Optional[float] = None
    bgVisible: Optional[bool] = None
    bgCalibration: Optional[dict] = None  # { p1x, p1y, p2x, p2y, knownFt }
    pxPerFt: Optional[float] = None
    widthFt: Optional[float] = None
    heightFt: Optional[float] = None


@api.patch("/ballrooms/{ballroom_id}/floor-plan")
async def set_floor_plan(ballroom_id: int, body: FloorPlanInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    sets, params = [], {"i": ballroom_id}
    if body.backgroundImageUrl is not None:
        sets.append("background_image_url=:url"); params["url"] = body.backgroundImageUrl
    if body.scaleFactor is not None:
        sets.append("scale_factor=:sf"); params["sf"] = body.scaleFactor
    if not sets: raise HTTPException(400, "Nothing to update")
    row = (await db.execute(text(f"UPDATE ballrooms SET {', '.join(sets)} WHERE id=:i RETURNING *"), params)).mappings().first()
    if not row: raise HTTPException(404, "Ballroom not found")
    await log_activity(db, user, "ballroom_floorplan_update", ballroom_id=ballroom_id)
    await db.commit()
    return ballroom_to_api(row)


@api.patch("/ballrooms/{ballroom_id}/canvas-settings")
async def set_canvas_settings(ballroom_id: int, body: CanvasSettingsInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    import json as _json
    mapping = {
        "snapEnabled": "snap_enabled", "gridSizeIn": "grid_size_in",
        "bgOpacity": "bg_opacity", "bgVisible": "bg_visible",
        "pxPerFt": "px_per_ft", "widthFt": "width_ft", "heightFt": "height_ft",
    }
    sets, params = [], {"i": ballroom_id}
    for k, col in mapping.items():
        v = getattr(body, k)
        if v is not None:
            sets.append(f"{col}=:{col}"); params[col] = v
    if body.bgCalibration is not None:
        sets.append("bg_calibration=CAST(:bgc AS jsonb)"); params["bgc"] = _json.dumps(body.bgCalibration)
    if not sets: raise HTTPException(400, "Nothing to update")
    row = (await db.execute(text(f"UPDATE ballrooms SET {', '.join(sets)} WHERE id=:i RETURNING *"), params)).mappings().first()
    if not row: raise HTTPException(404, "Ballroom not found")
    await log_activity(db, user, "ballroom_canvas_settings_update", ballroom_id=ballroom_id,
                       details=body.model_dump(exclude_none=True))
    await db.commit()
    return ballroom_to_api(row)


class CanvasObjectInput(BaseModel):
    ballroomId: int
    objectType: str  # 'stage' | 'dance_floor' | 'bar' | 'buffet' | 'carving' | 'pillar' | 'entrance' | 'exit' | 'blocker' | 'wall' | 'door' | 'room_ballroom' | 'room_bathroom' | 'room_hallway' | 'sign' | 'marker'
    label: Optional[str] = None
    x: float = 0
    y: float = 0
    width: float = 80
    height: float = 80
    rotation: float = 0
    properties: Optional[dict] = None


class CanvasObjectUpdate(BaseModel):
    label: Optional[str] = None
    x: Optional[float] = None
    y: Optional[float] = None
    width: Optional[float] = None
    height: Optional[float] = None
    rotation: Optional[float] = None
    properties: Optional[dict] = None


def canvas_obj_to_api(o) -> dict:
    pos = o["position"] or {}; dim = o["dimensions"] or {}
    props = o["properties"] if "properties" in o.keys() and o["properties"] is not None else {}
    return {"id": o["id"], "ballroomId": o["ballroom_id"], "objectType": o["object_type"],
            "label": o["label"],
            "x": float(pos.get("x", 0)), "y": float(pos.get("y", 0)),
            "width": float(dim.get("width", 80)), "height": float(dim.get("height", 80)),
            "rotation": float(o["rotation"] or 0),
            "properties": props}


@api.get("/ballrooms/{ballroom_id}/canvas-objects")
async def list_canvas_objects(ballroom_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text(
        "SELECT * FROM canvas_objects WHERE ballroom_id=:i ORDER BY id"
    ), {"i": ballroom_id})).mappings().all()
    return [canvas_obj_to_api(r) for r in rows]


@api.post("/canvas-objects", status_code=201)
async def create_canvas_object(body: CanvasObjectInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    import json as _json
    row = (await db.execute(text("""
        INSERT INTO canvas_objects (ballroom_id, object_type, label, position, dimensions, rotation, properties)
        VALUES (:b, :t, :l, CAST(:p AS jsonb), CAST(:d AS jsonb), :r, CAST(:pr AS jsonb)) RETURNING *
    """), {"b": body.ballroomId, "t": body.objectType, "l": body.label,
           "p": _json.dumps({"x": body.x, "y": body.y}),
           "d": _json.dumps({"width": body.width, "height": body.height}),
           "r": body.rotation,
           "pr": _json.dumps(body.properties or {})})).mappings().first()
    await log_activity(db, user, "canvas_object_create", ballroom_id=body.ballroomId, details={"type": body.objectType})
    await db.commit()
    return canvas_obj_to_api(row)


@api.patch("/canvas-objects/{obj_id}")
async def update_canvas_object(obj_id: int, body: CanvasObjectUpdate, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    import json as _json
    cur = (await db.execute(text("SELECT * FROM canvas_objects WHERE id=:i"), {"i": obj_id})).mappings().first()
    if not cur: raise HTTPException(404, "Object not found")
    pos = dict(cur["position"] or {})
    dim = dict(cur["dimensions"] or {})
    if body.x is not None: pos["x"] = body.x
    if body.y is not None: pos["y"] = body.y
    if body.width is not None: dim["width"] = body.width
    if body.height is not None: dim["height"] = body.height
    sets, params = ["position=CAST(:p AS jsonb)", "dimensions=CAST(:d AS jsonb)"], {
        "p": _json.dumps(pos), "d": _json.dumps(dim), "i": obj_id,
    }
    if body.label is not None: sets.append("label=:l"); params["l"] = body.label
    if body.rotation is not None: sets.append("rotation=:r"); params["r"] = body.rotation
    if body.properties is not None:
        # merge into existing
        existing = dict(cur["properties"] or {}) if "properties" in cur.keys() else {}
        existing.update(body.properties)
        sets.append("properties=CAST(:pr AS jsonb)"); params["pr"] = _json.dumps(existing)
    row = (await db.execute(text(f"UPDATE canvas_objects SET {', '.join(sets)} WHERE id=:i RETURNING *"), params)).mappings().first()
    await record_history(db, user, "canvas", "canvas_object_update", "canvas_object", str(obj_id),
                         _obj_snap(cur), _obj_snap(row))
    await db.commit()
    return canvas_obj_to_api(row)


@api.delete("/canvas-objects/{obj_id}", status_code=204)
async def delete_canvas_object(obj_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("SELECT ballroom_id FROM canvas_objects WHERE id=:i"), {"i": obj_id})).mappings().first()
    if not row: raise HTTPException(404, "Object not found")
    await db.execute(text("DELETE FROM canvas_objects WHERE id=:i"), {"i": obj_id})
    await log_activity(db, user, "canvas_object_delete", ballroom_id=row["ballroom_id"])
    await db.commit()


@api.get("/tables")
async def list_tables(
    user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
    ballroomId: Optional[int] = None,
):
    conds = []
    params = {}
    if ballroomId is not None:
        conds.append("t.ballroom_id = :br"); params["br"] = ballroomId
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows = (await db.execute(text(f"""
        SELECT t.*, COALESCE(SUM(g.party_size), 0) AS seated
        FROM tables t
        LEFT JOIN guests g ON g.table_id = t.id
        {where}
        GROUP BY t.id
        ORDER BY t.table_number
    """), params)).mappings().all()
    return [table_to_api(r, int(r["seated"])) for r in rows]


@api.post("/tables", status_code=201)
async def create_table(body: TableInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    # If typeId provided, hydrate defaults from the table type
    w_in, l_in, cap, shape = body.widthIn, body.lengthIn, body.maxCapacity, body.shape
    if body.typeId is not None:
        tt = (await db.execute(text("SELECT * FROM table_types WHERE id=:i AND is_active=TRUE"),
                                {"i": body.typeId})).mappings().first()
        if not tt: raise HTTPException(400, "Invalid or inactive table type")
        shape = tt["shape"]
        cap = body.maxCapacity if body.maxCapacity != 10 else int(tt["default_seats"])
        if w_in is None: w_in = float(tt["width_in"])
        if l_in is None: l_in = float(tt["length_in"])
    if w_in is None: w_in = 60.0 if shape == "round" else 96.0
    if l_in is None: l_in = 60.0 if shape == "round" else 48.0
    if shape == "square": l_in = w_in
    row = (await db.execute(text("""
        INSERT INTO tables (table_number, label, ballroom_id, shape, max_capacity,
                            canvas_x, canvas_y, rotation, notes, width_in, length_in, type_id, group_id)
        VALUES (:n, :l, :br, :sh, :c, :x, :y, :r, :nt, :wi, :li, :ti, :gi) RETURNING *
    """), {"n": body.tableNumber, "l": body.label, "br": body.ballroomId, "sh": shape,
           "c": cap, "x": body.canvasX or 0, "y": body.canvasY or 0,
           "r": body.rotation or 0, "nt": body.notes, "wi": w_in, "li": l_in,
           "ti": body.typeId, "gi": body.groupId})).mappings().first()
    await log_activity(db, user, "table_create", table_id=row["id"], ballroom_id=body.ballroomId,
                       details={"number": body.tableNumber, "capacity": cap, "typeId": body.typeId})
    await db.commit()
    return table_to_api(row, 0)


@api.patch("/tables/{table_id}")
async def update_table(table_id: int, body: TableUpdateInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    before = (await db.execute(text("SELECT * FROM tables WHERE id=:i"), {"i": table_id})).mappings().first()
    if not before: raise HTTPException(404, "Table not found")
    mapping = {"tableNumber": "table_number", "label": "label", "ballroomId": "ballroom_id",
               "shape": "shape", "maxCapacity": "max_capacity", "canvasX": "canvas_x",
               "canvasY": "canvas_y", "rotation": "rotation", "notes": "notes",
               "widthIn": "width_in", "lengthIn": "length_in",
               "typeId": "type_id", "groupId": "group_id"}
    sets, params = [], {"id": table_id}
    for k, col in mapping.items():
        v = getattr(body, k)
        if v is not None: sets.append(f"{col}=:{col}"); params[col] = v
    if not sets: raise HTTPException(400, "No fields to update")
    row = (await db.execute(text(f"UPDATE tables SET {', '.join(sets)} WHERE id=:id RETURNING *"), params)).mappings().first()
    seated = (await db.execute(text("SELECT COALESCE(SUM(party_size),0) FROM guests WHERE table_id=:i"), {"i": table_id})).scalar()
    await log_activity(db, user, "table_update", table_id=table_id, details=body.model_dump(exclude_none=True))
    await record_history(db, user, "canvas", "table_update", "table", str(table_id),
                         _table_snap(before), _table_snap(row))
    await db.commit()
    return table_to_api(row, int(seated))


@api.delete("/tables/{table_id}", status_code=204)
async def delete_table(table_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    seated = (await db.execute(text("SELECT COUNT(*) FROM guests WHERE table_id=:i"), {"i": table_id})).scalar()
    if seated:
        raise HTTPException(400, f"{seated} guests are seated at this table — unassign them first")
    await db.execute(text("DELETE FROM tables WHERE id=:i"), {"i": table_id})
    await log_activity(db, user, "table_delete", table_id=table_id)
    await db.commit()


@api.get("/tables/{table_id}")
async def get_table(table_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("SELECT * FROM tables WHERE id=:i"), {"i": table_id})).mappings().first()
    if not row: raise HTTPException(404, "Table not found")
    guests = (await db.execute(text("""
        SELECT g.*, sa.physically_seated, sa.assigned_at, sa.assigned_by_session AS assigned_by
        FROM guests g
        LEFT JOIN seat_assignments sa ON sa.guest_id = g.id AND sa.table_id = g.table_id
        WHERE g.table_id = :i
        ORDER BY g.full_name
    """), {"i": table_id})).mappings().all()
    seated = sum(g["party_size"] for g in guests)
    out = table_to_api(row, seated)
    out["guests"] = [{
        **guest_to_api(g),
        "physicallySeated": bool(g["physically_seated"]) if g["physically_seated"] is not None else False,
    } for g in guests]
    return out


@api.post("/tables/{table_id}/assign")
async def assign_to_table(table_id: int, body: AssignInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    t = (await db.execute(text("SELECT * FROM tables WHERE id=:i"), {"i": table_id})).mappings().first()
    if not t: raise HTTPException(404, "Table not found")
    g = (await db.execute(text("SELECT * FROM guests WHERE id=:i"), {"i": body.guestId})).mappings().first()
    if not g: raise HTTPException(404, "Guest not found")
    seated = (await db.execute(text("SELECT COALESCE(SUM(party_size),0) FROM guests WHERE table_id=:i AND id<>:gid"),
                                {"i": table_id, "gid": body.guestId})).scalar() or 0
    if seated + g["party_size"] > t["max_capacity"] and not body.allowOverflow:
        raise HTTPException(409, {
            "error": "capacity_exceeded",
            "message": f"Adding party of {g['party_size']} exceeds table capacity ({t['max_capacity']} seats, {seated} taken). Set allowOverflow=true to force.",
            "capacity": t["max_capacity"], "seatsTaken": seated, "partySize": g["party_size"],
        })

    # detect preference satisfaction
    other_ids = [r["id"] for r in (await db.execute(text("SELECT id FROM guests WHERE table_id=:i AND id<>:g"),
                                                     {"i": table_id, "g": body.guestId})).mappings().all()]
    pref_meta = {"mutualWith": [], "oneWayWith": []}
    if other_ids:
        rows = (await db.execute(text("""
            SELECT a.guest_id, a.resolved_guest_id,
                   (EXISTS(SELECT 1 FROM preference_resolutions b
                    WHERE b.guest_id=a.resolved_guest_id AND b.resolved_guest_id=a.guest_id AND b.resolution_status='confirmed')) AS mutual
            FROM preference_resolutions a
            WHERE a.resolution_status='confirmed'
              AND ((a.guest_id=:g AND a.resolved_guest_id = ANY(:others))
                OR (a.resolved_guest_id=:g AND a.guest_id = ANY(:others)))
        """), {"g": body.guestId, "others": other_ids})).mappings().all()
        for r in rows:
            other_id = r["resolved_guest_id"] if r["guest_id"] == body.guestId else r["guest_id"]
            (pref_meta["mutualWith"] if r["mutual"] else pref_meta["oneWayWith"]).append(other_id)

    await db.execute(text("""
        UPDATE guests SET table_id=:t, ballroom_id=:b, status='fully_assigned', last_updated_timestamp=NOW()
        WHERE id=:g
    """), {"t": table_id, "b": t["ballroom_id"], "g": body.guestId})

    await db.execute(text("DELETE FROM seat_assignments WHERE guest_id=:g"), {"g": body.guestId})
    await db.execute(text("""
        INSERT INTO seat_assignments (guest_id, table_id, assigned_by_session)
        VALUES (:g, :t, :s)
    """), {"g": body.guestId, "t": table_id, "s": user["display_name"]})

    await log_activity(db, user, "seat_assign", guest_id=body.guestId, table_id=table_id,
                       ballroom_id=t["ballroom_id"], details={"prefMeta": pref_meta, "overflow": body.allowOverflow})
    await db.commit()
    return {"ok": True, "preferenceMatch": pref_meta}


@api.post("/tables/{table_id}/unassign/{guest_id}", status_code=200)
async def unassign(table_id: int, guest_id: int, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(text("""
        UPDATE guests SET table_id=NULL, ballroom_id=NULL, status='unassigned', last_updated_timestamp=NOW()
        WHERE id=:g AND table_id=:t RETURNING id
    """), {"g": guest_id, "t": table_id})
    if not res.fetchone(): raise HTTPException(404, "Guest not at this table")
    await db.execute(text("DELETE FROM seat_assignments WHERE guest_id=:g AND table_id=:t"), {"g": guest_id, "t": table_id})
    await log_activity(db, user, "seat_unassign", guest_id=guest_id, table_id=table_id)
    await db.commit()
    return {"ok": True}


@api.patch("/tables/{table_id}/guests/{guest_id}/seated")
async def mark_seated(table_id: int, guest_id: int, body: SeatedInput, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(text("""
        UPDATE seat_assignments SET physically_seated=:s WHERE guest_id=:g AND table_id=:t RETURNING id
    """), {"s": body.seated, "g": guest_id, "t": table_id})
    if not res.fetchone(): raise HTTPException(404, "Assignment not found")
    await log_activity(db, user, "seat_check", guest_id=guest_id, table_id=table_id, details={"seated": body.seated})
    await db.commit()
    return {"ok": True}


@api.post("/seating/auto-suggest")
async def auto_suggest(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
                        ballroomId: Optional[int] = None):
    """Returns a non-binding plan: list of {guestId, tableId, ballroomId, reason}.
    Algorithm: build mutual-match clusters, sort by descending size, fit largest clusters first
    into tables that can hold them. Then place remaining single guests by descending party_size.
    """
    where_clause = "WHERE t.ballroom_id=:b" if ballroomId else ""
    params = {"b": ballroomId} if ballroomId else {}
    res = await db.execute(text(f"""
        SELECT t.*, COALESCE(SUM(g.party_size),0) AS seated
        FROM tables t LEFT JOIN guests g ON g.table_id = t.id
        {where_clause}
        GROUP BY t.id ORDER BY t.max_capacity
    """), params)
    tables = res.mappings().all() or []
    if not tables: raise HTTPException(400, "No tables available — create tables first")

    guests = (await db.execute(text(
        "SELECT id, full_name, party_size FROM guests WHERE status='unassigned'"
    ))).mappings().all()
    if not guests:
        return {"plan": [], "summary": "No unassigned guests"}

    # mutual cluster build via union-find on confirmed mutual edges
    parent = {g["id"]: g["id"] for g in guests}
    def find(x):
        while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[ra] = rb

    gid_set = {g["id"] for g in guests}
    mutuals = (await db.execute(text("""
        SELECT a.guest_id, a.resolved_guest_id FROM preference_resolutions a
        JOIN preference_resolutions b ON b.guest_id=a.resolved_guest_id AND b.resolved_guest_id=a.guest_id
        WHERE a.resolution_status='confirmed' AND b.resolution_status='confirmed' AND a.guest_id < a.resolved_guest_id
    """))).fetchall()
    for a, b in mutuals:
        if a in gid_set and b in gid_set: union(a, b)

    by_root = {}
    for g in guests:
        by_root.setdefault(find(g["id"]), []).append(g)
    clusters = sorted(by_root.values(), key=lambda c: -sum(x["party_size"] for x in c))

    remaining = [{"id": t["id"], "ballroom_id": t["ballroom_id"], "cap": t["max_capacity"],
                  "seated": int(t["seated"]), "number": t["table_number"]} for t in tables]
    plan = []
    for cluster in clusters:
        size = sum(x["party_size"] for x in cluster)
        # smallest table that fits the cluster
        slot = next((t for t in remaining if (t["cap"] - t["seated"]) >= size), None)
        if not slot:
            # fall back to largest with room
            slot = max(remaining, key=lambda t: t["cap"] - t["seated"])
            if slot["cap"] - slot["seated"] <= 0: continue
        for g in cluster:
            if slot["cap"] - slot["seated"] >= g["party_size"]:
                plan.append({"guestId": g["id"], "tableId": slot["id"],
                             "ballroomId": slot["ballroom_id"], "tableNumber": slot["number"],
                             "guestName": g["full_name"], "partySize": g["party_size"],
                             "reason": "mutual_cluster" if len(cluster) > 1 else "size_fit"})
                slot["seated"] += g["party_size"]

    return {"plan": plan, "summary": f"{len(plan)} of {len(guests)} unassigned guests can be seated"}


@api.post("/seating/auto-suggest/apply")
async def auto_suggest_apply(plan: List[dict], user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    applied = 0
    for item in plan:
        gid, tid = item.get("guestId"), item.get("tableId")
        if not gid or not tid: continue
        t = (await db.execute(text("SELECT ballroom_id, max_capacity FROM tables WHERE id=:i"), {"i": tid})).mappings().first()
        if not t: continue
        await db.execute(text("""
            UPDATE guests SET table_id=:t, ballroom_id=:b, status='fully_assigned', last_updated_timestamp=NOW() WHERE id=:g
        """), {"t": tid, "b": t["ballroom_id"], "g": gid})
        await db.execute(text("DELETE FROM seat_assignments WHERE guest_id=:g"), {"g": gid})
        await db.execute(text("""
            INSERT INTO seat_assignments (guest_id, table_id, assigned_by_session) VALUES (:g, :t, :s)
        """), {"g": gid, "t": tid, "s": user["display_name"]})
        applied += 1
    await log_activity(db, user, "auto_suggest_apply", details={"applied": applied, "total": len(plan)})
    await db.commit()
    return {"applied": applied}


# ---------- TABLE TYPES (inventory) ----------
class TableTypeInput(BaseModel):
    name: str = Field(min_length=1)
    shape: str = "round"
    defaultSeats: int = 10
    widthIn: float = 60
    lengthIn: float = 60
    quantityOwned: int = 0
    isActive: bool = True
    notes: Optional[str] = None


def tt_to_api(r) -> dict:
    return {"id": r["id"], "name": r["name"], "shape": r["shape"],
            "defaultSeats": int(r["default_seats"]),
            "widthIn": float(r["width_in"]), "lengthIn": float(r["length_in"]),
            "quantityOwned": int(r["quantity_owned"]), "isActive": bool(r["is_active"]),
            "notes": r["notes"], "createdAt": r["created_at"].isoformat()}


@api.get("/table-types")
async def list_table_types(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(text("SELECT * FROM table_types ORDER BY name"))).mappings().all()
    return [tt_to_api(r) for r in rows]


@api.post("/table-types", status_code=201)
async def create_table_type(body: TableTypeInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    shape = body.shape
    w_in = body.widthIn; l_in = body.lengthIn if shape != "square" else body.widthIn
    row = (await db.execute(text("""
        INSERT INTO table_types (name, shape, default_seats, width_in, length_in, quantity_owned, is_active, notes)
        VALUES (:n, :s, :ds, :w, :l, :q, :a, :nt) RETURNING *
    """), {"n": body.name, "s": shape, "ds": body.defaultSeats, "w": w_in, "l": l_in,
           "q": body.quantityOwned, "a": body.isActive, "nt": body.notes})).mappings().first()
    await log_activity(db, user, "table_type_create", details={"id": row["id"], "name": body.name})
    await db.commit()
    return tt_to_api(row)


@api.patch("/table-types/{tt_id}")
async def update_table_type(tt_id: int, body: TableTypeInput, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    shape = body.shape
    w_in = body.widthIn; l_in = body.lengthIn if shape != "square" else body.widthIn
    row = (await db.execute(text("""
        UPDATE table_types SET name=:n, shape=:s, default_seats=:ds, width_in=:w, length_in=:l,
            quantity_owned=:q, is_active=:a, notes=:nt
        WHERE id=:i RETURNING *
    """), {"i": tt_id, "n": body.name, "s": shape, "ds": body.defaultSeats, "w": w_in, "l": l_in,
           "q": body.quantityOwned, "a": body.isActive, "nt": body.notes})).mappings().first()
    if not row: raise HTTPException(404, "Table type not found")
    await log_activity(db, user, "table_type_update", details={"id": tt_id})
    await db.commit()
    return tt_to_api(row)


@api.delete("/table-types/{tt_id}", status_code=204)
async def delete_table_type(tt_id: int, user: dict = Depends(require_admin), db: AsyncSession = Depends(get_db)):
    in_use = (await db.execute(text("SELECT COUNT(*) FROM tables WHERE type_id=:i"), {"i": tt_id})).scalar()
    if in_use:
        raise HTTPException(400, f"Type is in use by {in_use} tables; deactivate instead")
    await db.execute(text("DELETE FROM table_types WHERE id=:i"), {"i": tt_id})
    await log_activity(db, user, "table_type_delete", details={"id": tt_id})
    await db.commit()


# ---------- BULK GUEST IMPORT (CSV / Excel / QuickBooks export) ----------
@app.post("/api/guests/bulk-import")
async def guests_bulk_import(
    request: Request,
    user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    """Accepts:
      - multipart upload with field 'file' (CSV, .xlsx, or .xls)
      - raw CSV text body when content-type is text/csv
    Columns (header row required, case-insensitive). Required: full_name (or name) AND
      either invoice_number (or invoice) OR a family_id.
    Optional columns: party_size (or guests), family_id, near_family_id,
      seating_preferences (semicolon-separated), high_chair_count, special_notes,
      email, phone.
    Auto-fills invoice from family_id (e.g. 'FAM-001-1') if missing.
    Upserts by invoice_number — existing rows updated, new rows inserted.
    """
    import csv as _csv, io
    ctype = (request.headers.get("content-type") or "").lower()
    rows_raw: List[dict] = []

    if "multipart" in ctype:
        form = await request.form()
        f = form.get("file")
        if f is None:
            raise HTTPException(400, "Missing 'file' field")
        fname = (getattr(f, "filename", "") or "").lower()
        data = await f.read()
        if fname.endswith((".xlsx", ".xls")):
            try:
                from openpyxl import load_workbook
            except Exception:
                raise HTTPException(500, "openpyxl not installed on server")
            wb = load_workbook(io.BytesIO(data), data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(it, [])]
            for r in it:
                row = {headers[i]: (str(r[i]).strip() if r[i] is not None else "") for i in range(min(len(headers), len(r)))}
                if any(v for v in row.values()): rows_raw.append(row)
        else:
            content = data.decode("utf-8-sig", errors="replace")
            reader = _csv.DictReader(io.StringIO(content))
            for row in reader: rows_raw.append({k: (v or "").strip() for k, v in row.items()})
    else:
        content = (await request.body()).decode("utf-8-sig", errors="replace")
        if not content.strip(): raise HTTPException(400, "Empty body")
        reader = _csv.DictReader(io.StringIO(content))
        for row in reader: rows_raw.append({k: (v or "").strip() for k, v in row.items()})

    if not rows_raw:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["No data rows found"]}

    norm = lambda s: (s or "").strip().lower().replace(" ", "_")
    sample_keys = {norm(k) for k in rows_raw[0].keys()}
    aliases = {
        "full_name": ["full_name", "name", "guest_name", "customer", "customer_name"],
        "invoice_number": ["invoice_number", "invoice", "invoiceno", "invoice#", "invoice_no", "invoice_num"],
        "party_size": ["party_size", "party", "guests", "guest_count", "headcount", "people"],
        "family_id": ["family_id", "family", "familyid"],
        "near_family_id": ["near_family_id", "near_family", "adjacent_family", "seat_near"],
        "preferences": ["seating_preferences", "preferences", "prefers"],
        "high_chair_count": ["high_chair_count", "high_chairs", "highchair"],
        "special_notes": ["special_notes", "notes", "note"],
    }
    def pick(row, key):
        for a in aliases[key]:
            for k, v in row.items():
                if norm(k) == a and (v or "").strip(): return v
        return None

    inserted = updated = skipped = 0
    errors: List[str] = []
    for i, row in enumerate(rows_raw, start=2):
        nm = (pick(row, "full_name") or "").strip()
        inv = (pick(row, "invoice_number") or "").strip()
        fam = (pick(row, "family_id") or "").strip() or None
        if not nm:
            skipped += 1; errors.append(f"Row {i}: missing name"); continue
        if not inv:
            # synthesize from family_id+row, or hash of name
            if fam:
                inv = f"{fam}-r{i}"
            else:
                inv = f"BULK-{nm.upper().replace(' ', '')[:12]}-r{i}"
        try:
            ps = int(pick(row, "party_size") or "1")
        except Exception:
            ps = 1
        try:
            hcc = int(pick(row, "high_chair_count") or "0")
        except Exception:
            hcc = 0
        near = (pick(row, "near_family_id") or "").strip() or None
        notes = pick(row, "special_notes")
        prefs_raw = pick(row, "preferences") or ""
        prefs = [p.strip() for p in prefs_raw.replace(",", ";").split(";") if p.strip()][:5]

        try:
            existing = (await db.execute(text("SELECT id FROM guests WHERE invoice_number=:i"),
                                          {"i": inv})).mappings().first()
            if existing:
                await db.execute(text("""
                    UPDATE guests SET full_name=:n, party_size=:ps, family_id=COALESCE(:fam, family_id),
                        near_family_id=COALESCE(:near, near_family_id),
                        seating_preferences=:prefs, high_chair_count=:hcc, high_chair_needed=:hcn,
                        special_notes=COALESCE(:nt, special_notes), last_updated_timestamp=NOW()
                    WHERE invoice_number=:i
                """), {"n": nm, "ps": ps, "fam": fam, "near": near, "prefs": prefs,
                       "hcc": hcc, "hcn": hcc > 0, "nt": notes, "i": inv})
                updated += 1
            else:
                await db.execute(text("""
                    INSERT INTO guests (full_name, invoice_number, party_size, seating_preferences,
                        high_chair_needed, high_chair_count, special_notes, family_id, near_family_id, is_duplicate)
                    VALUES (:n, :i, :ps, :prefs, :hcn, :hcc, :nt, :fam, :near, FALSE)
                """), {"n": nm, "i": inv, "ps": ps, "prefs": prefs,
                       "hcn": hcc > 0, "hcc": hcc, "nt": notes, "fam": fam, "near": near})
                inserted += 1
        except Exception as ex:
            skipped += 1; errors.append(f"Row {i}: {str(ex)[:140]}")

    await log_activity(db, user, "guests_bulk_import",
                       details={"inserted": inserted, "updated": updated, "skipped": skipped})
    await db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors[:30]}


# ---------- AUTOMATED SEATING ENGINE (family-grouped, combine-when-big) ----------
class AutoAssignBody(BaseModel):
    ballroomId: Optional[int] = None
    apply: bool = False  # if true, persist; otherwise return plan only
    allowCombine: bool = True  # combine multiple tables when a family > capacity


@api.post("/seating/auto-assign")
async def auto_assign(body: AutoAssignBody, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Group unassigned guests by family_id (singletons treated as a 1-family).
    For each family:
      - Try to find a single open table that fits the entire family. Prefer tables already
        seating near_family_id guests (adjacency boost).
      - If no single table fits AND allowCombine: link consecutive empty tables under a
        new group_id so the family is split across the minimum needed combined tables.
      - Fill tables to capacity (so few gray chairs).
    Returns a plan; persists when body.apply == true.
    """
    import uuid as _uuid

    where_t = "WHERE t.ballroom_id=:b" if body.ballroomId else ""
    params: dict = {}
    if body.ballroomId: params["b"] = body.ballroomId

    rows = (await db.execute(text(f"""
        SELECT t.*, COALESCE(SUM(g.party_size),0) AS seated
        FROM tables t LEFT JOIN guests g ON g.table_id = t.id
        {where_t}
        GROUP BY t.id ORDER BY t.table_number ASC
    """), params)).mappings().all()
    if not rows:
        raise HTTPException(400, "No tables — create or place tables on the canvas first")
    tables = [{
        "id": r["id"], "number": r["table_number"], "ballroom_id": r["ballroom_id"],
        "cap": int(r["max_capacity"]), "seated": int(r["seated"]),
        "x": float(r["canvas_x"] or 0), "y": float(r["canvas_y"] or 0),
        "group_id": r["group_id"] if "group_id" in r.keys() else None,
        "family_assigned": set(),  # family_ids on this table
    } for r in rows]

    g_where = "WHERE status='unassigned'"
    g_params: dict = {}
    if body.ballroomId:
        # only consider guests not already in another ballroom
        g_where += " AND (ballroom_id IS NULL OR ballroom_id = :br)"
        g_params["br"] = body.ballroomId
    guests = (await db.execute(text(
        f"SELECT id, full_name, party_size, family_id, near_family_id FROM guests {g_where}"
    ), g_params)).mappings().all()

    if not guests:
        return {"plan": [], "summary": "No unassigned guests", "applied": 0}

    # Already-assigned guests give us "family_assigned" anchors for adjacency
    anchored = (await db.execute(text("""
        SELECT table_id, family_id FROM guests
        WHERE table_id IS NOT NULL AND family_id IS NOT NULL
    """))).mappings().all()
    by_tid = {t["id"]: t for t in tables}
    for a in anchored:
        if a["table_id"] in by_tid and a["family_id"]:
            by_tid[a["table_id"]]["family_assigned"].add(a["family_id"])

    # Group unassigned by family_id (None → singleton synthetic family)
    families: dict = {}
    for g in guests:
        fid = g["family_id"] or f"__solo_{g['id']}"
        families.setdefault(fid, {
            "fid": fid, "near": g["near_family_id"],
            "members": [], "total": 0,
        })
        families[fid]["members"].append(dict(g))
        families[fid]["total"] += g["party_size"]

    # Sort families by descending total — bigger families need room first
    family_list = sorted(families.values(), key=lambda f: -f["total"])

    plan: List[dict] = []

    def score_table(t, family):
        """Higher = better. Free seats must >= 1 to be usable here for a partial fit."""
        free = t["cap"] - t["seated"]
        if free <= 0: return -1
        bonus = 0
        # adjacency boost
        if family["near"] and family["near"] in t["family_assigned"]:
            bonus += 1000
        # fill-to-capacity preference (smallest free that still fits)
        return bonus + (100 - free)  # smaller free = higher score

    for fam in family_list:
        remaining = fam["total"]
        # 1) find single table that fits everyone (skip tables already locked into a combined group)
        candidates = [t for t in tables
                      if (t["cap"] - t["seated"]) >= remaining
                      and not t.get("group_id")]
        candidates.sort(key=lambda t: -score_table(t, fam))
        if candidates:
            slot = candidates[0]
            for m in fam["members"]:
                plan.append({
                    "guestId": m["id"], "tableId": slot["id"],
                    "ballroomId": slot["ballroom_id"], "tableNumber": slot["number"],
                    "guestName": m["full_name"], "partySize": m["party_size"],
                    "familyId": fam["fid"] if not fam["fid"].startswith("__solo_") else None,
                    "groupId": slot["group_id"],
                    "reason": "fits_single_table",
                })
                slot["seated"] += m["party_size"]
            if not fam["fid"].startswith("__solo_"):
                slot["family_assigned"].add(fam["fid"])
            continue

        # 2) combine tables (multi-table group) if allowed
        if body.allowCombine:
            # rank tables by adjacency + free-seats descending — only ungrouped tables
            sortable = sorted(
                [t for t in tables
                 if (t["cap"] - t["seated"]) > 0 and not t.get("group_id")],
                key=lambda t: (
                    -(1 if (fam["near"] and fam["near"] in t["family_assigned"]) else 0),
                    -(t["cap"] - t["seated"]),
                ),
            )
            picked = []
            cum = 0
            for t in sortable:
                picked.append(t)
                cum += t["cap"] - t["seated"]
                if cum >= remaining: break
            if cum >= remaining and picked:
                gid = f"grp-{_uuid.uuid4().hex[:8]}"
                # Anchor = first picked table. Seat ALL members at the anchor (even if anchor
                # cap is exceeded) but link the other tables under the same group_id so the
                # UI shows them combined and accumulates the capacity.
                anchor = picked[0]
                for m in fam["members"]:
                    plan.append({
                        "guestId": m["id"], "tableId": anchor["id"],
                        "ballroomId": anchor["ballroom_id"], "tableNumber": anchor["number"],
                        "guestName": m["full_name"], "partySize": m["party_size"],
                        "familyId": fam["fid"] if not fam["fid"].startswith("__solo_") else None,
                        "groupId": gid,
                        "reason": "combined_tables",
                    })
                    anchor["seated"] += m["party_size"]
                for t in picked:
                    if not fam["fid"].startswith("__solo_"):
                        t["family_assigned"].add(fam["fid"])
                    t["group_id"] = gid
                continue

        # 3) cannot seat — leave unassigned, report
        plan.append({
            "guestId": None, "tableId": None,
            "guestName": ", ".join(m["full_name"] for m in fam["members"]),
            "partySize": fam["total"],
            "familyId": fam["fid"] if not fam["fid"].startswith("__solo_") else None,
            "reason": "no_capacity",
        })

    applied = 0
    if body.apply:
        # 1) write group_ids first (so tables show as combined even before guests move)
        gid_to_tables: dict = {}
        for t in tables:
            if t.get("group_id"):
                gid_to_tables.setdefault(t["group_id"], []).append(t["id"])
        for gid, ids in gid_to_tables.items():
            await db.execute(text("UPDATE tables SET group_id=:g WHERE id = ANY(:ids)"),
                             {"g": gid, "ids": ids})
        # 2) write guest assignments
        for item in plan:
            if not item.get("guestId") or not item.get("tableId"): continue
            await db.execute(text("""
                UPDATE guests SET table_id=:t, ballroom_id=:b, status='fully_assigned',
                    last_updated_timestamp=NOW() WHERE id=:g
            """), {"t": item["tableId"], "b": item["ballroomId"], "g": item["guestId"]})
            await db.execute(text("DELETE FROM seat_assignments WHERE guest_id=:g"),
                             {"g": item["guestId"]})
            await db.execute(text("""
                INSERT INTO seat_assignments (guest_id, table_id, assigned_by_session)
                VALUES (:g, :t, :s)
            """), {"g": item["guestId"], "t": item["tableId"], "s": user["display_name"]})
            applied += 1
        await log_activity(db, user, "auto_assign", details={"applied": applied, "total": len(plan)})
        await db.commit()

    unseated = sum(1 for p in plan if p["reason"] == "no_capacity")
    return {
        "plan": plan,
        "summary": f"{applied if body.apply else 0} guests seated · {unseated} family group(s) couldn't fit",
        "applied": applied,
        "unseatedFamilies": unseated,
    }


# ---------- FAMILY-MOVE (whole-family reassignment from guest list) ----------
class FamilyMoveBody(BaseModel):
    guestId: int                  # any guest in the family
    targetTableId: Optional[int]  # null = unassign


@api.post("/guests/family/move")
async def move_family(body: FamilyMoveBody, user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    g = (await db.execute(text("SELECT id, family_id FROM guests WHERE id=:i"),
                           {"i": body.guestId})).mappings().first()
    if not g: raise HTTPException(404, "Guest not found")
    fid = g["family_id"]
    if fid:
        members = (await db.execute(text("SELECT id, party_size FROM guests WHERE family_id=:f"),
                                     {"f": fid})).mappings().all()
    else:
        members = [g]

    if body.targetTableId:
        t = (await db.execute(text("SELECT ballroom_id, max_capacity FROM tables WHERE id=:i"),
                               {"i": body.targetTableId})).mappings().first()
        if not t: raise HTTPException(404, "Target table not found")
        # current seated AT TARGET (excluding family members we're moving in)
        member_ids = [m["id"] for m in members]
        cur = (await db.execute(text("""
            SELECT COALESCE(SUM(party_size),0) FROM guests
            WHERE table_id=:t AND id <> ALL(:ids)
        """), {"t": body.targetTableId, "ids": member_ids})).scalar() or 0
        need = sum(m["party_size"] for m in members)
        if cur + need > t["max_capacity"]:
            raise HTTPException(409, {
                "error": "capacity_exceeded",
                "message": f"Family of {need} won't fit (table has {t['max_capacity']-cur} seats free)",
                "capacity": t["max_capacity"], "free": t["max_capacity"]-cur, "needed": need,
            })
        await db.execute(text("""
            UPDATE guests SET table_id=:t, ballroom_id=:b, status='fully_assigned',
                last_updated_timestamp=NOW()
            WHERE id = ANY(:ids)
        """), {"t": body.targetTableId, "b": t["ballroom_id"], "ids": member_ids})
    else:
        member_ids = [m["id"] for m in members]
        await db.execute(text("""
            UPDATE guests SET table_id=NULL, ballroom_id=NULL, status='unassigned',
                last_updated_timestamp=NOW()
            WHERE id = ANY(:ids)
        """), {"ids": member_ids})

    await log_activity(db, user, "family_move", guest_id=body.guestId,
                       table_id=body.targetTableId,
                       details={"familyId": fid, "memberCount": len(members)})
    await db.commit()
    return {"ok": True, "movedCount": len(members), "familyId": fid}


# ---------- ACTION HISTORY (undo / redo) ----------
def _guest_snap(row) -> dict:
    if not row: return None
    return {
        "id": row["id"], "full_name": row["full_name"], "invoice_number": row["invoice_number"],
        "party_size": row["party_size"], "seating_preferences": list(row["seating_preferences"] or []),
        "high_chair_needed": row["high_chair_needed"], "high_chair_count": row["high_chair_count"],
        "status": row["status"], "ballroom_id": row["ballroom_id"], "table_id": row["table_id"],
        "special_notes": row["special_notes"], "is_duplicate": row["is_duplicate"],
        "family_id": row["family_id"] if "family_id" in row.keys() else None,
        "near_family_id": row["near_family_id"] if "near_family_id" in row.keys() else None,
    }

def _table_snap(row) -> dict:
    if not row: return None
    return {
        "id": row["id"], "table_number": row["table_number"], "label": row["label"],
        "ballroom_id": row["ballroom_id"], "shape": row["shape"], "max_capacity": row["max_capacity"],
        "canvas_x": float(row["canvas_x"] or 0), "canvas_y": float(row["canvas_y"] or 0),
        "rotation": float(row["rotation"] or 0),
        "width_in": float(row["width_in"]) if row["width_in"] is not None else None,
        "length_in": float(row["length_in"]) if row["length_in"] is not None else None,
        "notes": row["notes"],
        "type_id": row["type_id"] if "type_id" in row.keys() else None,
        "group_id": row["group_id"] if "group_id" in row.keys() else None,
    }

def _obj_snap(row) -> dict:
    if not row: return None
    pos = row["position"] or {}; dim = row["dimensions"] or {}
    return {
        "id": row["id"], "ballroom_id": row["ballroom_id"], "object_type": row["object_type"],
        "label": row["label"],
        "x": float(pos.get("x", 0)), "y": float(pos.get("y", 0)),
        "width": float(dim.get("width", 80)), "height": float(dim.get("height", 80)),
        "rotation": float(row["rotation"] or 0),
        "properties": row["properties"] or {},
    }

async def record_history(db: AsyncSession, user: dict, scope: str, action_type: str,
                         target_type: str, target_id: Optional[str],
                         before: Optional[dict], after: Optional[dict]) -> None:
    """Append an undo-able action to history. Truncates the redo stack (anything is_undone)."""
    try:
        sid = user["id"] if isinstance(user, dict) and "id" in user else user.get("id") if hasattr(user, "get") else None
        # whenever a NEW action is recorded, purge any redo-able (is_undone) actions further up the stack
        if sid is not None:
            await db.execute(text(
                "DELETE FROM action_history WHERE staff_id=:s AND is_undone=TRUE"
            ), {"s": sid})
        await db.execute(text("""
            INSERT INTO action_history (staff_id, scope, action_type, target_type, target_id,
                                        before_state, after_state)
            VALUES (:s, :sc, :a, :tt, :ti, CAST(:b AS jsonb), CAST(:af AS jsonb))
        """), {"s": sid, "sc": scope, "a": action_type, "tt": target_type, "ti": target_id,
                "b": json.dumps(before) if before is not None else None,
                "af": json.dumps(after) if after is not None else None})
    except Exception as ex:
        # never fail the main action because of history bookkeeping
        logger.warning(f"record_history failed: {ex}")


async def _apply_snap_guest(db: AsyncSession, snap: dict) -> None:
    if not snap: return
    await db.execute(text("""
        UPDATE guests SET full_name=:n, invoice_number=:i, party_size=:p,
            seating_preferences=:pf, high_chair_needed=:hcn, high_chair_count=:hcc,
            status=:st, ballroom_id=:b, table_id=:t, special_notes=:sn,
            is_duplicate=:dp, family_id=:fid, near_family_id=:nfid,
            last_updated_timestamp=NOW() WHERE id=:gid
    """), {"n": snap["full_name"], "i": snap["invoice_number"], "p": snap["party_size"],
            "pf": snap["seating_preferences"], "hcn": snap["high_chair_needed"],
            "hcc": snap["high_chair_count"], "st": snap["status"], "b": snap["ballroom_id"],
            "t": snap["table_id"], "sn": snap["special_notes"], "dp": snap["is_duplicate"],
            "fid": snap.get("family_id"), "nfid": snap.get("near_family_id"),
            "gid": snap["id"]})

async def _apply_snap_table(db: AsyncSession, snap: dict) -> None:
    if not snap: return
    await db.execute(text("""
        UPDATE tables SET table_number=:n, label=:l, ballroom_id=:b, shape=:s,
            max_capacity=:c, canvas_x=:x, canvas_y=:y, rotation=:r, width_in=:wi,
            length_in=:li, notes=:nt, type_id=:ti, group_id=:gi
        WHERE id=:id
    """), {"id": snap["id"], "n": snap["table_number"], "l": snap["label"],
            "b": snap["ballroom_id"], "s": snap["shape"], "c": snap["max_capacity"],
            "x": snap["canvas_x"], "y": snap["canvas_y"], "r": snap["rotation"],
            "wi": snap["width_in"], "li": snap["length_in"], "nt": snap["notes"],
            "ti": snap.get("type_id"), "gi": snap.get("group_id")})

async def _apply_snap_object(db: AsyncSession, snap: dict) -> None:
    if not snap: return
    await db.execute(text("""
        UPDATE canvas_objects SET label=:l, position=CAST(:p AS jsonb),
            dimensions=CAST(:d AS jsonb), rotation=:r, properties=CAST(:pr AS jsonb)
        WHERE id=:id
    """), {"id": snap["id"], "l": snap["label"],
            "p": json.dumps({"x": snap["x"], "y": snap["y"]}),
            "d": json.dumps({"width": snap["width"], "height": snap["height"]}),
            "r": snap["rotation"],
            "pr": json.dumps(snap["properties"] or {})})

async def _apply_history_state(db: AsyncSession, action_row, *, redo: bool) -> dict:
    """Apply before_state (undo) or after_state (redo) of an action."""
    snap = action_row["after_state"] if redo else action_row["before_state"]
    if snap is None and not redo:
        # action created the row → undo means delete
        await _delete_target(db, action_row["target_type"], action_row["target_id"])
        return {"deleted": action_row["target_type"], "id": action_row["target_id"]}
    tt = action_row["target_type"]
    if tt == "guest":   await _apply_snap_guest(db, snap)
    elif tt == "table": await _apply_snap_table(db, snap)
    elif tt == "canvas_object": await _apply_snap_object(db, snap)
    return {"applied": tt, "id": action_row["target_id"]}

async def _delete_target(db: AsyncSession, target_type: str, target_id: str) -> None:
    if not target_id: return
    tbl = {"guest": "guests", "table": "tables", "canvas_object": "canvas_objects"}.get(target_type)
    if not tbl: return
    try:
        await db.execute(text(f"DELETE FROM {tbl} WHERE id=:i"), {"i": int(target_id)})
    except Exception:
        pass


@api.post("/history/undo")
async def history_undo(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("""
        SELECT * FROM action_history WHERE staff_id=:s AND is_undone=FALSE
        ORDER BY created_at DESC LIMIT 1
    """), {"s": user["id"]})).mappings().first()
    if not row:
        return {"ok": False, "reason": "no_actions"}
    res = await _apply_history_state(db, row, redo=False)
    await db.execute(text("UPDATE action_history SET is_undone=TRUE WHERE id=:i"), {"i": row["id"]})
    await db.commit()
    return {"ok": True, "actionId": row["id"], "actionType": row["action_type"],
            "scope": row["scope"], "targetType": row["target_type"], "result": res}


@api.post("/history/redo")
async def history_redo(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    row = (await db.execute(text("""
        SELECT * FROM action_history WHERE staff_id=:s AND is_undone=TRUE
        ORDER BY created_at ASC LIMIT 1
    """), {"s": user["id"]})).mappings().first()
    if not row:
        return {"ok": False, "reason": "no_actions"}
    res = await _apply_history_state(db, row, redo=True)
    await db.execute(text("UPDATE action_history SET is_undone=FALSE WHERE id=:i"), {"i": row["id"]})
    await db.commit()
    return {"ok": True, "actionId": row["id"], "actionType": row["action_type"],
            "scope": row["scope"], "targetType": row["target_type"], "result": res}


@api.get("/history/stack")
async def history_stack(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    undo_count = (await db.execute(text(
        "SELECT COUNT(*) FROM action_history WHERE staff_id=:s AND is_undone=FALSE"
    ), {"s": user["id"]})).scalar()
    redo_count = (await db.execute(text(
        "SELECT COUNT(*) FROM action_history WHERE staff_id=:s AND is_undone=TRUE"
    ), {"s": user["id"]})).scalar()
    return {"undoAvailable": int(undo_count or 0), "redoAvailable": int(redo_count or 0)}


# ---------- CONFLICT DETECTION ----------
@api.get("/seating/conflicts")
async def list_conflicts(ballroomId: Optional[int] = None,
                          user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Return seating conflicts grouped per table.
    Types:
      - over_capacity: seatsTaken > effective capacity (account for group_id)
      - family_split: same family_id seated at multiple tables (without group)
      - near_family_not_at_same_table: family's near_family_id has a guest seated elsewhere
      - one_way_preference: guest A lists guest B in seatingPreferences but B doesn't reciprocate
    """
    where_g = "WHERE g.table_id IS NOT NULL"
    params: dict = {}
    if ballroomId:
        where_g += " AND g.ballroom_id=:b"; params["b"] = ballroomId
    guests = (await db.execute(text(f"""
        SELECT g.*, t.table_number, t.max_capacity, t.group_id AS t_group_id
        FROM guests g JOIN tables t ON t.id=g.table_id
        {where_g}
    """), params)).mappings().all()

    tables_q = "SELECT * FROM tables"
    if ballroomId: tables_q += " WHERE ballroom_id=:b"
    tables = (await db.execute(text(tables_q), params)).mappings().all()

    by_tid = {t["id"]: dict(t) for t in tables}
    seated_per_t: dict = {}
    by_family: dict = {}
    name_index: dict = {}
    for g in guests:
        seated_per_t.setdefault(g["table_id"], 0)
        seated_per_t[g["table_id"]] += int(g["party_size"] or 0)
        if g["family_id"]:
            by_family.setdefault(g["family_id"], []).append(dict(g))
        name_index[g["full_name"].lower()] = dict(g)
    # group capacity
    group_cap: dict = {}
    for t in tables:
        if t["group_id"]:
            group_cap[t["group_id"]] = group_cap.get(t["group_id"], 0) + int(t["max_capacity"] or 0)

    conflicts: List[dict] = []
    # over capacity
    for tid, seated in seated_per_t.items():
        t = by_tid.get(tid)
        if not t: continue
        eff = group_cap.get(t["group_id"]) if t["group_id"] else int(t["max_capacity"] or 0)
        if seated > eff:
            conflicts.append({
                "type": "over_capacity", "tableId": tid,
                "tableNumber": t["table_number"], "seated": seated, "capacity": eff,
                "message": f"Table {t['table_number']} has {seated} guests but seats {eff}",
            })
    # family split (members at different tables AND tables don't share a group_id)
    for fid, members in by_family.items():
        tids = {m["table_id"] for m in members}
        if len(tids) > 1:
            # check if all those tables share the same group_id
            gids = {by_tid[tid]["group_id"] for tid in tids if tid in by_tid}
            if not (len(gids) == 1 and None not in gids):
                conflicts.append({
                    "type": "family_split", "familyId": fid,
                    "tableIds": list(tids),
                    "memberCount": sum(int(m["party_size"] or 0) for m in members),
                    "message": f"Family {fid} is split across {len(tids)} tables",
                })
    # near_family adjacency missing
    for g in guests:
        nf = g["near_family_id"]
        if not nf: continue
        peers = by_family.get(nf, [])
        if not peers: continue
        my_tid = g["table_id"]
        if not any(p["table_id"] == my_tid or
                   (by_tid.get(p["table_id"], {}).get("group_id") and
                    by_tid.get(p["table_id"], {}).get("group_id") == by_tid.get(my_tid, {}).get("group_id"))
                   for p in peers):
            conflicts.append({
                "type": "near_family_not_at_same_table",
                "guestId": g["id"], "familyId": g["family_id"], "wantedFamilyId": nf,
                "tableId": my_tid, "tableNumber": g["table_number"],
                "message": f"{g['full_name']} wanted to sit near family {nf}",
            })
    # one-way preferences
    for g in guests:
        for pref in (g["seating_preferences"] or []):
            target = name_index.get((pref or "").strip().lower())
            if not target: continue
            theirs = [p.strip().lower() for p in (target["seating_preferences"] or [])]
            if g["full_name"].lower() not in theirs:
                conflicts.append({
                    "type": "one_way_preference",
                    "guestId": g["id"], "guestName": g["full_name"],
                    "targetGuestId": target["id"], "targetGuestName": target["full_name"],
                    "tableId": g["table_id"], "tableNumber": g["table_number"],
                    "message": f"{g['full_name']} requested {target['full_name']} but not vice versa",
                })

    # group by table
    by_table_id: dict = {}
    for c in conflicts:
        tid = c.get("tableId") or (c.get("tableIds")[0] if c.get("tableIds") else None)
        if tid:
            by_table_id.setdefault(tid, []).append(c)
    return {"conflicts": conflicts, "byTableId": by_table_id, "count": len(conflicts)}


# ---------- ANALYTICS ----------
@api.get("/analytics/summary")
async def analytics_summary(user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    stats = (await db.execute(text("""
        SELECT
          COUNT(*) AS submissions,
          COALESCE(SUM(party_size),0) AS people,
          COALESCE(SUM(CASE WHEN status='fully_assigned' THEN party_size ELSE 0 END),0) AS seated_people,
          COUNT(*) FILTER (WHERE status='fully_assigned') AS seated,
          COUNT(*) FILTER (WHERE status='partially_assigned') AS partial,
          COUNT(*) FILTER (WHERE status='unassigned') AS unassigned,
          COALESCE(SUM(high_chair_count),0) AS high_chairs
        FROM guests
    """))).mappings().first()
    t_stats = (await db.execute(text("""
        SELECT COUNT(*) AS table_count, COALESCE(SUM(max_capacity),0) AS total_cap
        FROM tables
    """))).mappings().first()
    seated_capacity = (await db.execute(text("""
        SELECT COALESCE(SUM(g.party_size),0) FROM guests g
        WHERE g.table_id IS NOT NULL
    """))).scalar() or 0
    total_cap = int(t_stats["total_cap"] or 0)
    utilization = round(100 * seated_capacity / total_cap, 1) if total_cap > 0 else 0.0
    conflicts = await list_conflicts(None, user, db)
    return {
        "totalSubmissions": int(stats["submissions"] or 0),
        "totalPeople": int(stats["people"] or 0),
        "seatedSubmissions": int(stats["seated"] or 0),
        "seatedPeople": int(stats["seated_people"] or 0),
        "unassignedSubmissions": int(stats["unassigned"] or 0),
        "partialSubmissions": int(stats["partial"] or 0),
        "tableCount": int(t_stats["table_count"] or 0),
        "totalCapacity": total_cap,
        "tableUtilizationPct": utilization,
        "highChairsRequested": int(stats["high_chairs"] or 0),
        "activeConflicts": conflicts["count"],
    }


# ---------- DEV / SKIP-LOGIN ----------
@api.post("/auth/dev-login")
async def dev_login(db: AsyncSession = Depends(get_db)):
    """Test-only skip-login. Disabled unless DEV_AUTH_BYPASS=1 in backend/.env.
    Returns a real JWT for the admin user — no password required.
    """
    if os.environ.get("DEV_AUTH_BYPASS") != "1":
        raise HTTPException(403, "Dev login is disabled in this environment")
    row = (await db.execute(text(
        "SELECT * FROM staff_users WHERE is_admin=TRUE AND is_active=TRUE ORDER BY id LIMIT 1"
    ))).mappings().first()
    if not row: raise HTTPException(500, "No admin user available")
    await db.execute(text("UPDATE staff_users SET last_login=NOW() WHERE id=:i"), {"i": row["id"]})
    await log_activity(db, dict(row), "dev_login")
    await db.commit()
    token = make_token(row["id"], row["username"], row["is_admin"])
    return {"token": token, "user": {"id": row["id"], "username": row["username"],
            "displayName": row["display_name"], "isAdmin": row["is_admin"]},
            "devBypass": True}


# ---------- ACTIVITY LOG ----------
@api.get("/activity-log")
async def activity_log(
    user: dict = Depends(get_current_user), db: AsyncSession = Depends(get_db),
    limit: int = 200, offset: int = 0, staffName: Optional[str] = None, actionType: Optional[str] = None,
):
    conds, params = [], {"limit": min(limit, 500), "offset": offset}
    if staffName: conds.append("staff_member_name ILIKE :sn"); params["sn"] = f"%{staffName}%"
    if actionType: conds.append("action_type = :at"); params["at"] = actionType
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows = (await db.execute(text(f"""
        SELECT id, action_type, staff_member_name, guest_id, table_id, ballroom_id, details, created_at
        FROM activity_log {where} ORDER BY created_at DESC LIMIT :limit OFFSET :offset
    """), params)).mappings().all()
    return [{"id": r["id"], "actionType": r["action_type"], "staffMemberName": r["staff_member_name"],
             "guestId": r["guest_id"], "tableId": r["table_id"], "ballroomId": r["ballroom_id"],
             "details": r["details"], "createdAt": r["created_at"].isoformat()} for r in rows]


# ---------- STARTUP ----------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=False,
    allow_methods=["*"], allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await init_db()
    async with AsyncSessionLocal() as db:
        admin_username = os.environ["ADMIN_USERNAME"]
        admin_password = os.environ["ADMIN_PASSWORD"]
        admin_display = os.environ["ADMIN_DISPLAY_NAME"]
        existing = (await db.execute(text("SELECT id, password_hash FROM staff_users WHERE username=:u"), {"u": admin_username})).mappings().first()
        if not existing:
            await db.execute(text("""
                INSERT INTO staff_users (username, password_hash, display_name, is_admin)
                VALUES (:u, :p, :d, TRUE)
            """), {"u": admin_username, "p": hash_pw(admin_password), "d": admin_display})
            await db.commit()
            logger.info(f"Seeded admin user: {admin_username}")
        elif not verify_pw(admin_password, existing["password_hash"]):
            await db.execute(text("UPDATE staff_users SET password_hash=:p WHERE id=:i"),
                             {"p": hash_pw(admin_password), "i": existing["id"]})
            await db.commit()
            logger.info(f"Updated admin password: {admin_username}")
